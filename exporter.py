"""
Excel exporter — produces a multi-sheet breakdown workbook.

Sheet 1: Scene Breakdown (script order) — the main working document
Sheet 2: By Location — scenes grouped and totalled by location
Sheet 3: Tally — recommendation summary with page counts
"""

from io import BytesIO
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Colour palette ───────────────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_SUBHEAD_BG  = "2F5496"
C_ALT_ROW     = "EBF0FA"

# Recommendation colours
_REC_COLOURS = {
    "VP (INT. VEHICLE)":  "E2EFDA",   # green
    "VP (INT. AIRCRAFT)": "D9EAD3",   # green-teal
    "LUMOSTAGE":          "FFF2CC",   # yellow
    "ON LOCATION":        "DDEBF7",   # light blue
    "OPTION: EITHER":     "F2F2F2",   # light grey
    "SET BUILD":          "FCE4D6",   # orange
    "VFX":                "F4CCCC",   # red-pink
    "OMITTED":            "D9D9D9",   # mid grey
}

def _rec_fill(recommendation) -> PatternFill:
    if not isinstance(recommendation, str):
        return PatternFill("solid", fgColor="FFFFFF")
    key = recommendation.upper().strip()
    for k, colour in _REC_COLOURS.items():
        if k in key:
            return PatternFill("solid", fgColor=colour)
    return PatternFill("solid", fgColor="FFFFFF")


# ── Helpers ──────────────────────────────────────────────────────────────────

def _hdr(ws, row: int, col: int, value: str, sub: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(
        name="Calibri", bold=True,
        color=C_HEADER_FG, size=9 if sub else 10,
    )
    cell.fill = PatternFill("solid", fgColor=C_SUBHEAD_BG if sub else C_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _thin_border() -> Border:
    s = Side(style="thin", color="C9C9C9")
    return Border(left=s, right=s, top=s, bottom=s)


def _set_col_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _eighths_to_float(s: str) -> float:
    """Convert '1 3/8' or '3/8' or '2' back to float pages."""
    s = s.strip()
    if not s:
        return 0.0
    parts = s.split()
    if len(parts) == 2:
        whole = int(parts[0])
        num, den = parts[1].split("/")
        return whole + int(num) / int(den)
    if "/" in parts[0]:
        num, den = parts[0].split("/")
        return int(num) / int(den)
    return float(parts[0])


# ── Main export ──────────────────────────────────────────────────────────────

def export_to_excel(scenes: list, script_title: str = "Screenplay") -> BytesIO:
    wb = openpyxl.Workbook()

    _build_scene_sheet(wb.active, scenes, script_title)
    _build_location_sheet(wb.create_sheet("By Location"), scenes)
    _build_tally_sheet(wb.create_sheet("Tally"), scenes)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Sheet 1: Scene Breakdown ─────────────────────────────────────────────────

_SCENE_HEADERS = [
    "Sc #", "SLUGLINE", "INT/EXT", "LOCATION", "TIME OF DAY",
    "PAGES", "APPROACH", "CONFIDENCE", "VFX NOTES", "PRODUCTION NOTES",
]
_SCENE_WIDTHS = [6, 42, 8, 30, 14, 7, 22, 14, 28, 32]


def _build_scene_sheet(ws, scenes: list, title: str) -> None:
    ws.title = "Scene Breakdown"

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_SCENE_HEADERS))
    title_cell = ws.cell(row=1, column=1, value=f"{title} — Scene Breakdown")
    title_cell.font = Font(name="Calibri", bold=True, color=C_HEADER_FG, size=13)
    title_cell.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row
    for col, h in enumerate(_SCENE_HEADERS, 1):
        _hdr(ws, 2, col, h)
    ws.row_dimensions[2].height = 22

    # Data rows
    for i, scene in enumerate(scenes, 3):
        rec = getattr(scene, "recommendation", "")
        conf = getattr(scene, "confidence", "")
        fill = _rec_fill(rec)

        values = [
            scene.number or str(i - 2),
            scene.raw_slug,
            scene.int_ext,
            scene.location,
            scene.time_of_day,
            scene.page_count_str,
            rec,
            conf,
            getattr(scene, "vfx_notes", "") or "",
            getattr(scene, "production_notes", "") or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = fill
            cell.border = _thin_border()
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=col in (2, 4, 9, 10),
            )

    _set_col_widths(ws, _SCENE_WIDTHS)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(_SCENE_HEADERS))}2"


# ── Sheet 2: By Location ─────────────────────────────────────────────────────

def _build_location_sheet(ws, scenes: list) -> None:
    ws.title = "By Location"

    # Group scenes by root location (before the first sub-location dash)
    groups: dict[str, list] = defaultdict(list)
    for s in scenes:
        root = s.location.split(" - ")[0].strip()
        groups[root].append(s)

    headers = ["LOCATION", "SCENE #", "SLUGLINE", "PAGES", "APPROACH", "NOTES"]
    widths  = [32, 8, 42, 7, 22, 32]

    for col, h in enumerate(headers, 1):
        _hdr(ws, 1, col, h)
    ws.row_dimensions[1].height = 22

    row = 2
    for root_loc in sorted(groups.keys()):
        loc_scenes = groups[root_loc]
        total_pages = sum(_eighths_to_float(s.page_count_str) for s in loc_scenes)
        total_str = _eighths_str_from_float(total_pages)

        # Location group header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        group_cell = ws.cell(
            row=row, column=1,
            value=f"{root_loc}  ({len(loc_scenes)} scenes · {total_str} pages)",
        )
        group_cell.font = Font(name="Calibri", bold=True, color=C_HEADER_FG, size=10)
        group_cell.fill = PatternFill("solid", fgColor=C_SUBHEAD_BG)
        group_cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

        for s in loc_scenes:
            rec = getattr(s, "recommendation", "")
            fill = _rec_fill(rec)
            for col, val in enumerate([
                s.location, s.number, s.raw_slug, s.page_count_str, rec, ""
            ], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = fill
                cell.border = _thin_border()
                cell.font = Font(name="Calibri", size=9)
                cell.alignment = Alignment(vertical="top", wrap_text=col in (3, 6))
            row += 1

        row += 1  # blank separator row

    _set_col_widths(ws, widths)
    ws.freeze_panes = "A2"


# ── Sheet 3: Tally ───────────────────────────────────────────────────────────

def _build_tally_sheet(ws, scenes: list) -> None:
    ws.title = "Tally"

    from collections import Counter
    rec_pages: dict[str, float]  = defaultdict(float)
    rec_scenes: dict[str, int]   = Counter()
    intex_pages: dict[str, float] = defaultdict(float)

    for s in scenes:
        rec_raw = getattr(s, "recommendation", "")
        rec = rec_raw if isinstance(rec_raw, str) and rec_raw else "Unassigned"
        p = _eighths_to_float(s.page_count_str)
        rec_pages[rec]   += p
        rec_scenes[rec]  += 1
        intex_pages[s.int_ext] += p

    total_pages = sum(rec_pages.values())

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12

    r = 1
    _hdr(ws, r, 1, "APPROACH"); _hdr(ws, r, 2, "SCENES"); _hdr(ws, r, 3, "PAGES"); _hdr(ws, r, 4, "% SCRIPT")
    ws.row_dimensions[r].height = 20
    r += 1

    for rec in sorted(rec_pages.keys()):
        pages = rec_pages[rec]
        pct = (pages / total_pages * 100) if total_pages else 0
        fill = _rec_fill(rec)
        for col, val in enumerate([rec, rec_scenes[rec], _eighths_str_from_float(pages), f"{pct:.1f}%"], 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = fill
            cell.border = _thin_border()
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
        r += 1

    # Totals row
    r += 1
    for col, val in enumerate(["TOTAL", sum(rec_scenes.values()), _eighths_str_from_float(total_pages), "100%"], 1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = Font(name="Calibri", bold=True, size=10, color=C_HEADER_FG)
        cell.fill = PatternFill("solid", fgColor=C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")

    r += 2
    # INT vs EXT breakdown
    _hdr(ws, r, 1, "INT vs EXT"); _hdr(ws, r, 2, "PAGES"); _hdr(ws, r, 3, "% SCRIPT")
    r += 1
    for ie, pages in sorted(intex_pages.items()):
        pct = (pages / total_pages * 100) if total_pages else 0
        ws.cell(row=r, column=1, value=ie).font = Font(name="Calibri", size=9)
        ws.cell(row=r, column=2, value=_eighths_str_from_float(pages)).font = Font(name="Calibri", size=9)
        ws.cell(row=r, column=3, value=f"{pct:.1f}%").font = Font(name="Calibri", size=9)
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = _thin_border()
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="center" if col > 1 else "left")
        r += 1


# ── Import from Excel ────────────────────────────────────────────────────────

def import_from_excel(xlsx_bytes: bytes) -> tuple[dict[str, dict], str]:
    """
    Read a previously exported Scene Breakdown xlsx.

    Returns:
        (data, error_msg) where data maps scene_number -> {recommendation,
        vfx_notes, production_notes} and error_msg is "" on success.

    Matching is by Sc # (column 1 of the Scene Breakdown sheet).  The function
    is tolerant of the title row, header row, and blank separator rows.
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)

        # Find the Scene Breakdown sheet (exact or fuzzy match)
        ws = None
        for name in wb.sheetnames:
            if "scene" in name.lower():
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active

        # Locate the header row by scanning the first 5 rows for a "Sc #" cell
        header_row_idx: int | None = None
        col_map: dict[str, int] = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
            for ci, cell_val in enumerate(row):
                if cell_val is None:
                    continue
                k = str(cell_val).strip().upper()
                if k in ("SC #", "SC#", "SCENE #", "SCENE#"):
                    header_row_idx = row_idx
                    break
            if header_row_idx:
                # Map column names → zero-based index
                row_vals = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
                for ci, cell_val in enumerate(row_vals):
                    if cell_val is None:
                        continue
                    k = str(cell_val).strip().upper()
                    if k in ("SC #", "SC#", "SCENE #", "SCENE#"):
                        col_map["number"] = ci
                    elif k in ("APPROACH", "RECOMMENDATION"):
                        col_map["recommendation"] = ci
                    elif "VFX" in k:
                        col_map["vfx_notes"] = ci
                    elif "PRODUCTION" in k:
                        col_map["production_notes"] = ci
                break

        wb.close()

        if header_row_idx is None or "number" not in col_map:
            return {}, "Could not find a Scene Breakdown header row (expected 'Sc #' column)."

        # Re-open for data (read_only mode doesn't support row slicing easily after close)
        wb2 = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
        ws2 = None
        for name in wb2.sheetnames:
            if "scene" in name.lower():
                ws2 = wb2[name]
                break
        if ws2 is None:
            ws2 = wb2.active

        result: dict[str, dict] = {}
        for row in ws2.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or all(c is None for c in row):
                continue
            raw_num = row[col_map["number"]]
            if raw_num is None:
                continue
            scene_num = str(raw_num).strip().rstrip(".0")  # handle Excel int→float like "5.0"
            if not scene_num:
                continue

            def _cell(key: str) -> str:
                idx = col_map.get(key)
                if idx is None or idx >= len(row):
                    return ""
                val = row[idx]
                return str(val).strip() if val is not None else ""

            result[scene_num] = {
                "recommendation":   _cell("recommendation"),
                "vfx_notes":        _cell("vfx_notes"),
                "production_notes": _cell("production_notes"),
            }

        wb2.close()
        return result, ""

    except Exception as exc:
        return {}, f"Failed to read xlsx: {exc}"


def apply_excel_import(scenes: list, data: dict[str, dict], valid_options: list[str]) -> int:
    """
    Apply data from import_from_excel() to a scenes list.

    Only updates a field if the imported value is non-empty.  Approach is
    only applied if it matches a known option (guards against typos or
    stale option names from an old export).

    Returns the count of scenes updated.
    """
    updated = 0
    scene_map = {s.number: s for s in scenes if s.number}
    for num, row in data.items():
        scene = scene_map.get(num)
        if scene is None:
            continue
        changed = False
        rec = row.get("recommendation", "")
        if rec and rec in valid_options:
            scene.recommendation = rec
            changed = True
        vfx = row.get("vfx_notes", "")
        if vfx:
            scene.vfx_notes = vfx
            changed = True
        notes = row.get("production_notes", "")
        if notes:
            scene.production_notes = notes
            changed = True
        if changed:
            updated += 1
    return updated


# ── Utility ──────────────────────────────────────────────────────────────────

def _eighths_str_from_float(p: float) -> str:
    total = round(p * 8)
    total = max(total, 0)
    whole, rem = divmod(total, 8)
    if whole > 0 and rem > 0:
        return f"{whole} {rem}/8"
    if whole > 0:
        return str(whole)
    if rem > 0:
        return f"{rem}/8"
    return "0"
